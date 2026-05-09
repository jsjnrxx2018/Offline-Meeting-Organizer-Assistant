from __future__ import annotations

import copy
from datetime import date, timedelta
import json
import os
import re
import sys
from dataclasses import dataclass, asdict
from pathlib import Path
from tkinter import BOTH, END, LEFT, RIGHT, TOP, X, Button, Entry, Frame, Label, StringVar, Text, Tk, filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, Side
try:
    import pythoncom
    import win32com.client
except ImportError:
    pythoncom = None
    win32com = None


APP_TITLE = "会议信息录入助手"
APP_DIR = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent
DEFAULT_EXCEL = APP_DIR / "会议安排.xlsx"
APP_DATA_DIR = Path(os.getenv("LOCALAPPDATA") or Path.home()) / "MeetingAssistant"
CONFIG_FILE = APP_DATA_DIR / "config.json"
FIELDS = [
    ("meeting_time", "会议时间"),
    ("meeting_name", "会议名称"),
    ("location", "会议地点"),
    ("department", "负责部门"),
    ("attendees", "出席范围"),
    ("convener", "召集人"),
    ("leaders", "出席领导"),
]

PROMPT_TEMPLATE = """提示词：

# Role
你是一名专业的行政秘书，擅长从非结构化的会议通知文本中提取关键信息，并将其整理为标准的格式化文本。

# Task
请阅读用户提供的【会议原始信息】，提取关键要素，并严格按照【输出格式】输出。

# Rules & Constraints
1. **时间处理**：
   - 如果原文是相对时间（如“下周二”），请根据当前日期（{{current_date}}）推算具体日期。
   - 如果原文已有具体日期（如“5/10”），请保持原样或标准化格式（如“5月10日”）。
   - 最终格式统一为：`M月D日 星期X HH:MM`。
2. **字段映射**：
   - `与会者` / `参会人员` -> 映射为 `出席范围`。
   - `召集部门` / `主办部门` -> 映射为 `负责部门`。
   - `出席领导` / `参会领导` -> 映射为 `出席领导`。
3. **空值处理（重要）**：
   - 如果原文中**未提及**某个字段的信息，该字段冒号后必须**留空**（即直接换行），**严禁**填入“无”、“未知”或“暂无”。
4. **内容清洗**：
   - 去除人名或部门名中多余的空格（除非是分隔符）。
   - 保持原文的专有名词（如会议名称）不变。

# Output Format
会议时间：{提取的日期和时间}
会议名称：{提取的会议全称}
会议地点：{提取的地点}
负责部门：{提取的负责部门，若无则留空}
出席范围: {提取的与会者}
召集人：{提取的召集人}
出席领导：{提取的出席领导，若无则留空}

# Input Data
{{用户输入的会议文本}}"""


def load_last_excel_path() -> str:
    try:
        data = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    except Exception:
        return str(DEFAULT_EXCEL)

    path = Path(str(data.get("excel_path", "")))
    if path.exists():
        return str(path)
    return str(DEFAULT_EXCEL)


def save_last_excel_path(path: str) -> None:
    try:
        APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
        CONFIG_FILE.write_text(json.dumps({"excel_path": path}, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


@dataclass
class Meeting:
    meeting_time: str = ""
    meeting_name: str = ""
    location: str = ""
    department: str = ""
    attendees: str = ""
    convener: str = ""
    leaders: str = ""

    @property
    def date_part(self) -> str:
        date, _ = split_datetime(self.meeting_time)
        return date

    @property
    def time_part(self) -> str:
        _, time = split_datetime(self.meeting_time)
        return time


def normalize_label(label: str) -> str:
    return re.sub(r"[\s:：\n]", "", label)


def chinese_number_to_int(text: str) -> int | None:
    text = text.strip()
    if not text:
        return None
    if text.isdigit():
        return int(text)

    digits = {"零": 0, "〇": 0, "一": 1, "二": 2, "两": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9}
    if text in digits:
        return digits[text]
    if text == "十":
        return 10
    if text.startswith("十"):
        tail = text[1:]
        return 10 + digits.get(tail, 0)
    if "十" in text:
        head, tail = text.split("十", 1)
        return digits.get(head, 0) * 10 + digits.get(tail, 0)
    return None


def weekday_name(index: int) -> str:
    return "一二三四五六日"[index]


def relative_date_text(text: str, today: date | None = None) -> str:
    today = today or date.today()
    weekday_map = {"一": 0, "二": 1, "三": 2, "四": 3, "五": 4, "六": 5, "日": 6, "天": 6}

    day_offsets = {"今天": 0, "明天": 1, "后天": 2}
    for word, offset in day_offsets.items():
        if word in text:
            target = today + timedelta(days=offset)
            return f"{target.month}月{target.day}日 星期{weekday_name(target.weekday())}"

    match = re.search(r"(下周|下星期|本周|这周|本星期|这星期|周|星期)([一二三四五六日天])", text)
    if not match:
        return ""

    prefix, weekday_text = match.groups()
    target_weekday = weekday_map[weekday_text]
    if prefix in ("下周", "下星期"):
        days = (7 - today.weekday()) + target_weekday
    else:
        days = target_weekday - today.weekday()
        if days < 0:
            days += 7
    target = today + timedelta(days=days)
    return f"{target.month}月{target.day}日 星期{weekday_name(target.weekday())}"


def natural_time_text(text: str) -> str:
    match = re.search(r"(上午|下午|晚上)?\s*([0-9]{1,2}|[一二两三四五六七八九十]{1,3})\s*[点时:：]\s*([0-9]{1,2}|[一二三四五六七八九十]{1,3})?", text)
    if not match:
        return ""
    period, hour_text, minute_text = match.groups()
    hour = chinese_number_to_int(hour_text)
    minute = chinese_number_to_int(minute_text or "0")
    if hour is None or minute is None:
        return ""
    if period in ("下午", "晚上") and hour < 12:
        hour += 12
    return f"{hour:02d}:{minute:02d}"


def natural_datetime_text(text: str) -> str:
    date_text = relative_date_text(text)
    time_text = natural_time_text(text)
    return f"{date_text} {time_text}".strip()


def split_datetime(value: str) -> tuple[str, str]:
    value = value.strip()
    if not value:
        return "", ""

    time_match = re.search(r"(?:(上午|下午|晚上)\s*)?([01]?\d|2[0-3]|[一二两三四五六七八九十]{1,3})[:：点时]([0-5]?\d|[一二三四五六七八九十]{1,3})?", value)
    time_text = ""
    if time_match:
        period, hour_text, minute_text = time_match.groups()
        hour = chinese_number_to_int(hour_text) or 0
        minute = chinese_number_to_int(minute_text or "00") or 0
        if period in ("下午", "晚上") and hour < 12:
            hour += 12
        time_text = f"{hour:02d}:{minute:02d}"

    date_text = value
    if time_match:
        date_text = (value[: time_match.start()] + value[time_match.end() :]).strip(" ，,;；")

    return date_text.strip(), time_text


def extract_labeled_fields(text: str) -> dict[str, str]:
    aliases = {
        "meeting_time": ["会议时间", "时间", "日期时间"],
        "meeting_name": ["会议名称", "会议全称", "名称", "会议内容"],
        "location": ["会议地点", "地点"],
        "department": ["负责部门", "承办部门", "责任部门"],
        "attendees": ["出席范围", "参会人员", "与会者", "参会范围", "参加人员"],
        "convener": ["召集人", "主持人"],
        "leaders": ["出席领导", "参会领导"],
    }
    label_to_field = {
        normalize_label(alias): field
        for field, names in aliases.items()
        for alias in names
    }

    result: dict[str, str] = {}
    pattern = re.compile(r"^\s*([^:：\n]{2,12})\s*[:：]\s*(.*?)\s*$")
    for line in text.splitlines():
        match = pattern.match(line)
        if not match:
            continue
        label, value = match.groups()
        field = label_to_field.get(normalize_label(label))
        if field:
            result[field] = value.strip()
    return result


def fallback_extract(text: str, result: dict[str, str]) -> None:
    compact = re.sub(r"\s+", " ", text).strip()
    if not compact:
        return

    if not result.get("meeting_time"):
        natural_datetime = natural_datetime_text(compact)
        if natural_datetime:
            result["meeting_time"] = natural_datetime

    if not result.get("meeting_time"):
        date_patterns = [
            r"\d{4}年\d{1,2}月\d{1,2}日(?:\s*[（(]?[周星期][一二三四五六日天][）)]?)?(?:\s*(?:上午|下午|晚上)?\s*\d{1,2}[:：点时]\d{0,2})?",
            r"\d{1,2}月\d{1,2}日(?:\s*[（(]?[周星期][一二三四五六日天][）)]?)?(?:\s*(?:上午|下午|晚上)?\s*\d{1,2}[:：点时]\d{0,2})?",
            r"\d{4}[-/]\d{1,2}[-/]\d{1,2}\s*\d{0,2}[:：]?\d{0,2}",
        ]
        for pattern in date_patterns:
            match = re.search(pattern, compact)
            if match:
                result["meeting_time"] = match.group(0).strip()
                break

    if not result.get("meeting_name"):
        name_match = re.search(r"(?:召开|参加|举行|出席)([^，。；;]{2,60}?(?:会议|会))", compact)
        if name_match:
            result["meeting_name"] = name_match.group(1).strip()
        else:
            quoted = re.search(r"[《“\"]([^》”\"]{2,60}?(?:会议|会))[》”\"]", compact)
            if quoted:
                result["meeting_name"] = quoted.group(1).strip()

    if not result.get("location"):
        loc_match = re.search(r"在([^，。；;]{2,40}?)(?:召开|举行|参加|出席)", compact)
        if not loc_match:
            loc_match = re.search(r"(?:在|地点为|地点[:：]?)([^，。；;]{2,40}?(?:会议室|报告厅|礼堂|中心|楼|室|厅))", compact)
        if loc_match:
            result["location"] = loc_match.group(1).strip()

    if not result.get("attendees"):
        attendees_match = re.search(r"(?:与会者|参会人员|出席范围)\s*[:：]?\s*([^，。；;]+)", compact)
        if attendees_match:
            result["attendees"] = re.sub(r"\s+", "、", attendees_match.group(1).strip())

    if not result.get("convener"):
        convener_match = re.search(r"召集人\s*(?:是|为|[:：])\s*([^，。；;]+)", compact)
        if convener_match:
            result["convener"] = convener_match.group(1).strip()


def parse_meeting(text: str) -> Meeting:
    values = extract_labeled_fields(text)
    fallback_extract(text, values)
    return Meeting(**{field: values.get(field, "") for field, _ in FIELDS})


def split_meeting_blocks(text: str) -> list[str]:
    blocks: list[list[str]] = []
    current: list[str] = []
    meeting_time_label = re.compile(r"^\s*会议时间\s*[:：]")
    saw_meeting_time = False

    for line in text.splitlines():
        if meeting_time_label.match(line):
            saw_meeting_time = True
            if current:
                blocks.append(current)
            current = [line]
        elif current:
            current.append(line)

    if current:
        blocks.append(current)

    if saw_meeting_time:
        return ["\n".join(block).strip() for block in blocks if "\n".join(block).strip()]
    return [text.strip()] if text.strip() else []


def parse_meetings(text: str) -> list[Meeting]:
    blocks = split_meeting_blocks(text)
    return [parse_meeting(block) for block in blocks] if blocks else []


def find_header_row(ws) -> int:
    expected = {"日期", "开始时间", "会议名称和内容", "地点", "负责部门", "出席范围", "召集人", "出席领导"}
    best_row = 1
    best_hits = 0
    for row in range(1, min(ws.max_row, 20) + 1):
        labels = {normalize_label(str(ws.cell(row, col).value or "")) for col in range(1, min(ws.max_column, 12) + 1)}
        hits = sum(1 for name in expected if normalize_label(name) in labels)
        if hits > best_hits:
            best_hits = hits
            best_row = row
    return best_row


def is_writable_row(ws, row: int) -> bool:
    return not any(isinstance(ws.cell(row, col), MergedCell) for col in range(1, 9))


def next_target_row(ws, header_row: int) -> int:
    for row in range(header_row + 1, ws.max_row + 1):
        if is_writable_row(ws, row) and all(ws.cell(row, col).value in (None, "") for col in range(1, 9)):
            return row
    row = ws.max_row + 1
    while not is_writable_row(ws, row):
        row += 1
    return row


def style_source_row(ws, header_row: int, target_row: int) -> int | None:
    for row in range(target_row - 1, header_row, -1):
        if is_writable_row(ws, row):
            return row
    if is_writable_row(ws, header_row + 1):
        return header_row + 1
    return None


def unmerge_date_cells(ws, header_row: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_col != 1 or merged_range.max_col != 1:
            continue
        if merged_range.max_row <= header_row:
            continue

        start_row = max(merged_range.min_row, header_row + 1)
        end_row = merged_range.max_row
        date_value = ws.cell(merged_range.min_row, 1).value
        ws.unmerge_cells(str(merged_range))
        for row in range(start_row, end_row + 1):
            if any(ws.cell(row, col).value not in (None, "") for col in range(2, 9)):
                ws.cell(row, 1).value = date_value


def parse_sort_date(value: object) -> tuple[int, int, int]:
    if isinstance(value, date):
        return value.year, value.month, value.day
    text = str(value or "").strip()
    if not text:
        return 9999, 12, 31

    match = re.search(r"(?:(\d{4})年)?\s*(\d{1,2})月\s*(\d{1,2})日", text)
    if match:
        year = int(match.group(1) or 9999)
        return year, int(match.group(2)), int(match.group(3))

    match = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if match:
        return int(match.group(1)), int(match.group(2)), int(match.group(3))

    return 9999, 12, 31


def normalize_date_display(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"(日)\s*(星期[一二三四五六日天])", r"\1\n\2", text)
    text = re.sub(r"(日)\s*(周[一二三四五六日天])", r"\1\n\2", text)
    return text


def parse_sort_time(value: object) -> tuple[int, int]:
    text = str(value or "").strip()
    match = re.search(r"([01]?\d|2[0-3])[:：点时]([0-5]?\d)?", text)
    if match:
        return int(match.group(1)), int(match.group(2) or "0")
    return 23, 59


def data_rows(ws, header_row: int) -> list[int]:
    rows: list[int] = []
    for row in range(header_row + 1, ws.max_row + 1):
        if not is_writable_row(ws, row):
            continue
        if any(ws.cell(row, col).value not in (None, "") for col in range(1, 9)):
            rows.append(row)
    return rows


def date_font(ws, header_row: int) -> Font:
    return copy.copy(ws.cell(header_row, 1).font)


def data_font(ws, header_row: int) -> Font:
    return copy.copy(ws.cell(header_row, 1).font)


def format_date_cell(ws, row: int, font: Font | None = None) -> None:
    cell = ws.cell(row, 1)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if font is not None:
        cell.font = copy.copy(font)


def table_border(ws) -> Border:
    header_border = ws.cell(find_header_row(ws), 1).border
    side = header_border.left if header_border.left and header_border.left.style else Side(style="double", color="000000")
    return Border(left=copy.copy(side), right=copy.copy(side), top=copy.copy(side), bottom=copy.copy(side))


def apply_date_group_style(ws, start_row: int, end_row: int, border: Border, font: Font) -> None:
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row, 1)
        cell.border = copy.copy(border)
        format_date_cell(ws, row, font)


def format_data_cell(ws, row: int, col: int, border: Border, font: Font) -> None:
    cell = ws.cell(row, col)
    cell.font = copy.copy(font)
    cell.border = copy.copy(border)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def format_data_rows(ws, header_row: int) -> None:
    rows = data_rows(ws, header_row)
    if not rows:
        return

    border = table_border(ws)
    font = data_font(ws, header_row)
    for row in rows:
        ws.row_dimensions[row].height = ws.row_dimensions[header_row].height
        for col in range(1, 9):
            if not isinstance(ws.cell(row, col), MergedCell):
                format_data_cell(ws, row, col, border, font)


def merge_same_dates(ws, header_row: int) -> None:
    rows = data_rows(ws, header_row)
    if not rows:
        return

    border = table_border(ws)
    font = date_font(ws, header_row)
    for row in rows:
        ws.cell(row, 1).value = normalize_date_display(ws.cell(row, 1).value)
        ws.cell(row, 1).border = copy.copy(border)
        format_date_cell(ws, row, font)

    group_start = rows[0]
    previous_date = parse_sort_date(ws.cell(group_start, 1).value)

    def merge_group(start_row: int, end_row: int) -> None:
        ws.cell(start_row, 1).value = normalize_date_display(ws.cell(start_row, 1).value)
        apply_date_group_style(ws, start_row, end_row, border, font)
        format_date_cell(ws, start_row, font)
        if end_row <= start_row:
            return
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        apply_date_group_style(ws, start_row, end_row, border, font)
        format_date_cell(ws, start_row, font)

    for row in rows[1:]:
        current_date = parse_sort_date(ws.cell(row, 1).value)
        if current_date != previous_date:
            merge_group(group_start, row - 1)
            group_start = row
            previous_date = current_date

    merge_group(group_start, rows[-1])


def sort_meetings(ws, header_row: int, track_row: int | None = None) -> int | None:
    unmerge_date_cells(ws, header_row)
    rows = data_rows(ws, header_row)
    if not rows:
        return None

    records = []
    for row in rows:
        values = [ws.cell(row, col).value for col in range(1, 9)]
        values[0] = normalize_date_display(values[0])
        records.append((parse_sort_date(values[0]), parse_sort_time(values[1]), row, values))

    records.sort(key=lambda item: (item[0], item[1], item[2]))
    first_row = rows[0]
    template_row = rows[0]

    tracked_target_row = None
    for offset, (_, _, original_row, values) in enumerate(records):
        target_row = first_row + offset
        if original_row == track_row:
            tracked_target_row = target_row
        if is_writable_row(ws, original_row):
            copy_row_style(ws, original_row, target_row)
        else:
            copy_row_style(ws, template_row, target_row)
        for col, value in enumerate(values, start=1):
            ws.cell(target_row, col).value = value

    for row in range(first_row + len(records), ws.max_row + 1):
        if is_writable_row(ws, row):
            for col in range(1, 9):
                ws.cell(row, col).value = None

    format_data_rows(ws, header_row)
    merge_same_dates(ws, header_row)
    return tracked_target_row


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, 9):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.font:
            target.font = copy.copy(source.font)
        if source.border:
            target.border = copy.copy(source.border)
        if source.fill:
            target.fill = copy.copy(source.fill)


def write_meeting_to_next_row(ws, header_row: int, meeting: Meeting) -> int:
    row = next_target_row(ws, header_row)

    style_row = style_source_row(ws, header_row, row)
    if style_row is not None:
        copy_row_style(ws, style_row, row)

    values = [
        normalize_date_display(meeting.date_part),
        meeting.time_part,
        meeting.meeting_name,
        meeting.location,
        meeting.department,
        meeting.attendees,
        meeting.convener,
        meeting.leaders,
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(row, col).value = value
    return row


def append_meetings(excel_path: Path, meetings: list[Meeting]) -> tuple[int, int | None]:
    wb = load_workbook(excel_path)
    ws = wb.active
    header_row = find_header_row(ws)
    unmerge_date_cells(ws, header_row)

    last_row = None
    for meeting in meetings:
        last_row = write_meeting_to_next_row(ws, header_row, meeting)

    sorted_row = sort_meetings(ws, header_row, last_row)
    wb.save(excel_path)
    return len(meetings), sorted_row


def excel_app():
    if win32com is None:
        raise RuntimeError("缺少 pywin32，无法实时操作已打开的 Excel。")
    pythoncom.CoInitialize()
    try:
        app = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        app = win32com.client.Dispatch("Excel.Application")
    app.Visible = True
    return app


def opened_workbook(app, excel_path: Path):
    target = str(excel_path.resolve()).lower()
    for workbook in app.Workbooks:
        try:
            if str(workbook.FullName).lower() == target:
                return workbook
        except Exception:
            continue
    raise RuntimeError("请先在 Excel 中打开所选工作簿，再点击添加会议。")


def com_cell_value(ws, row: int, col: int):
    return ws.Cells(row, col).Value


def com_find_header_row(ws) -> int:
    expected = {"日期", "开始时间", "会议名称和内容", "地点", "负责部门", "出席范围", "召集人", "出席领导"}
    best_row = 1
    best_hits = 0
    used_last = ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1
    for row in range(1, min(used_last, 20) + 1):
        labels = {normalize_label(str(com_cell_value(ws, row, col) or "")) for col in range(1, 9)}
        hits = sum(1 for name in expected if normalize_label(name) in labels)
        if hits > best_hits:
            best_hits = hits
            best_row = row
    return best_row


def com_unmerge_date_cells(ws, header_row: int) -> None:
    used_last = ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1
    seen: set[str] = set()
    for row in range(header_row + 1, used_last + 1):
        cell = ws.Cells(row, 1)
        if not cell.MergeCells:
            continue
        area = cell.MergeArea
        address = area.Address
        if address in seen:
            continue
        seen.add(address)
        if area.Column != 1 or area.Columns.Count != 1:
            continue
        value = area.Cells(1, 1).Value
        start_row = max(area.Row, header_row + 1)
        end_row = area.Row + area.Rows.Count - 1
        area.UnMerge()
        for fill_row in range(start_row, end_row + 1):
            if any(com_cell_value(ws, fill_row, col) not in (None, "") for col in range(2, 9)):
                ws.Cells(fill_row, 1).Value = value


def com_data_rows(ws, header_row: int) -> list[int]:
    used_last = max(ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1, header_row)
    rows: list[int] = []
    for row in range(header_row + 1, used_last + 1):
        if any(com_cell_value(ws, row, col) not in (None, "") for col in range(1, 9)):
            rows.append(row)
    return rows


def com_next_target_row(ws, header_row: int) -> int:
    used_last = max(ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1, header_row)
    for row in range(header_row + 1, used_last + 1):
        if all(com_cell_value(ws, row, col) in (None, "") for col in range(1, 9)):
            return row
    return used_last + 1


def com_write_meeting(ws, header_row: int, meeting: Meeting) -> int:
    row = com_next_target_row(ws, header_row)
    values = [
        normalize_date_display(meeting.date_part),
        meeting.time_part,
        meeting.meeting_name,
        meeting.location,
        meeting.department,
        meeting.attendees,
        meeting.convener,
        meeting.leaders,
    ]
    for col, value in enumerate(values, start=1):
        ws.Cells(row, col).Value = value
    return row


def com_format_range(ws, start_row: int, end_row: int, header_row: int) -> None:
    if end_row < start_row:
        return
    data_range = ws.Range(ws.Cells(start_row, 1), ws.Cells(end_row, 8))
    data_range.Font.Name = ws.Cells(header_row, 1).Font.Name
    data_range.Font.Size = ws.Cells(header_row, 1).Font.Size
    data_range.Font.Bold = ws.Cells(header_row, 1).Font.Bold
    data_range.HorizontalAlignment = -4108
    data_range.VerticalAlignment = -4108
    data_range.WrapText = True
    for border_index in (7, 8, 9, 10, 11, 12):
        border = data_range.Borders(border_index)
        border.LineStyle = -4119
        border.Weight = 4
    ws.Rows(f"{start_row}:{end_row}").RowHeight = ws.Rows(header_row).RowHeight


def com_merge_same_dates(ws, header_row: int) -> None:
    rows = com_data_rows(ws, header_row)
    if not rows:
        return
    for row in rows:
        ws.Cells(row, 1).Value = normalize_date_display(ws.Cells(row, 1).Value)

    group_start = rows[0]
    previous_date = parse_sort_date(ws.Cells(group_start, 1).Value)
    for row in rows[1:]:
        current_date = parse_sort_date(ws.Cells(row, 1).Value)
        if current_date != previous_date:
            if row - 1 > group_start:
                ws.Range(ws.Cells(group_start, 1), ws.Cells(row - 1, 1)).Merge()
            previous_date = current_date
            group_start = row
    if rows[-1] > group_start:
        ws.Range(ws.Cells(group_start, 1), ws.Cells(rows[-1], 1)).Merge()
    ws.Range(ws.Cells(rows[0], 1), ws.Cells(rows[-1], 1)).HorizontalAlignment = -4108
    ws.Range(ws.Cells(rows[0], 1), ws.Cells(rows[-1], 1)).VerticalAlignment = -4108
    ws.Range(ws.Cells(rows[0], 1), ws.Cells(rows[-1], 1)).WrapText = True


def com_sort_and_format(ws, header_row: int, track_row: int | None = None) -> int | None:
    com_unmerge_date_cells(ws, header_row)
    rows = com_data_rows(ws, header_row)
    if not rows:
        return None

    records = []
    for row in rows:
        values = [com_cell_value(ws, row, col) for col in range(1, 9)]
        values[0] = normalize_date_display(values[0])
        records.append((parse_sort_date(values[0]), parse_sort_time(values[1]), row, values))
    records.sort(key=lambda item: (item[0], item[1], item[2]))

    first_row = rows[0]
    last_row = first_row + len(records) - 1
    tracked_target_row = None
    for offset, (_, _, original_row, values) in enumerate(records):
        target_row = first_row + offset
        if original_row == track_row:
            tracked_target_row = target_row
        for col, value in enumerate(values, start=1):
            ws.Cells(target_row, col).Value = value

    used_last = ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1
    if used_last > last_row:
        ws.Range(ws.Cells(last_row + 1, 1), ws.Cells(used_last, 8)).ClearContents()

    com_format_range(ws, first_row, last_row, header_row)
    com_merge_same_dates(ws, header_row)
    return tracked_target_row


def append_meetings_live(excel_path: Path, meetings: list[Meeting]) -> tuple[int, int | None]:
    app = excel_app()
    workbook = opened_workbook(app, excel_path)
    worksheet = workbook.ActiveSheet
    header_row = com_find_header_row(worksheet)
    com_unmerge_date_cells(worksheet, header_row)

    last_row = None
    for meeting in meetings:
        last_row = com_write_meeting(worksheet, header_row, meeting)

    sorted_row = com_sort_and_format(worksheet, header_row, last_row)
    workbook.Save()
    app.Calculate()
    return len(meetings), sorted_row


def append_meeting(excel_path: Path, meeting: Meeting) -> int:
    _, row = append_meetings(excel_path, [meeting])
    return row or 0


def meeting_to_text(meeting: Meeting) -> str:
    values = asdict(meeting)
    return "\n".join(f"{label}：{values[field]}" for field, label in FIELDS)


def meetings_to_text(meetings: list[Meeting]) -> str:
    return "\n\n".join(f"【会议 {index}】\n{meeting_to_text(meeting)}" for index, meeting in enumerate(meetings, start=1))


class MeetingAssistantApp:
    def __init__(self, root: Tk) -> None:
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1480x720")
        self.excel_path = StringVar(value=load_last_excel_path())

        self.build_ui()

    def build_ui(self) -> None:
        file_frame = Frame(self.root, padx=12, pady=10)
        file_frame.pack(side=TOP, fill=X)
        Label(file_frame, text="Excel 文件：").pack(side=LEFT)
        Entry(file_frame, textvariable=self.excel_path).pack(side=LEFT, fill=X, expand=True, padx=(4, 8))
        Button(file_frame, text="选择文件", command=self.choose_file).pack(side=RIGHT)

        main = Frame(self.root, padx=12, pady=4)
        main.pack(fill=BOTH, expand=True)
        main.grid_columnconfigure(0, weight=3, uniform="columns")
        main.grid_columnconfigure(1, weight=3, uniform="columns")
        main.grid_columnconfigure(2, weight=4, uniform="columns")
        main.grid_rowconfigure(0, weight=1)

        left = Frame(main)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        Label(left, text="会议信息输入").pack(anchor="w")
        self.input_text = Text(left, wrap="word", height=24)
        self.input_text.pack(fill=BOTH, expand=True, pady=(4, 0))

        right = Frame(main)
        right.grid(row=0, column=1, sticky="nsew", padx=(8, 8))
        Label(right, text="解析结果预览").pack(anchor="w")
        self.preview_text = Text(right, wrap="word", height=24, state="disabled")
        self.preview_text.pack(fill=BOTH, expand=True, pady=(4, 0))

        template = Frame(main)
        template.grid(row=0, column=2, sticky="nsew", padx=(8, 0))
        Label(template, text="提示词模板").pack(anchor="w")
        self.template_text = Text(template, wrap="word", height=24)
        self.template_text.pack(fill=BOTH, expand=True, pady=(4, 0))
        self.template_text.insert("1.0", PROMPT_TEMPLATE)
        self.template_text.configure(state="disabled")

        buttons = Frame(self.root, padx=12, pady=12)
        buttons.pack(fill=X)
        Button(buttons, text="解析预览", width=16, command=self.preview).pack(side=LEFT)
        Button(buttons, text="添加会议", width=16, command=self.add_meeting).pack(side=LEFT, padx=10)
        Button(buttons, text="清空", width=16, command=self.clear).pack(side=LEFT)

    def choose_file(self) -> None:
        path = filedialog.askopenfilename(
            title="选择会议安排 Excel 文件",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
            initialdir=str(Path(self.excel_path.get()).parent),
        )
        if path:
            self.excel_path.set(path)
            save_last_excel_path(path)

    def get_input(self) -> str:
        return self.input_text.get("1.0", END).strip()

    def preview(self) -> list[Meeting] | None:
        text = self.get_input()
        if not text:
            messagebox.showwarning(APP_TITLE, "请先输入会议信息。")
            return None
        meetings = parse_meetings(text)
        if not meetings:
            messagebox.showwarning(APP_TITLE, "没有解析到会议信息。")
            return None
        self.preview_text.configure(state="normal")
        self.preview_text.delete("1.0", END)
        self.preview_text.insert("1.0", meetings_to_text(meetings))
        self.preview_text.configure(state="disabled")
        return meetings

    def add_meeting(self) -> None:
        meetings = self.preview()
        if meetings is None:
            return
        incomplete = [index for index, meeting in enumerate(meetings, start=1) if not meeting.meeting_time or not meeting.meeting_name]
        if incomplete:
            joined = "、".join(str(index) for index in incomplete)
            if not messagebox.askyesno(APP_TITLE, f"第 {joined} 条会议的会议时间或会议名称为空，仍然添加吗？"):
                return

        excel_path = Path(self.excel_path.get())
        if not excel_path.exists():
            messagebox.showerror(APP_TITLE, f"Excel 文件不存在：\n{excel_path}")
            return
        save_last_excel_path(str(excel_path))

        try:
            count, row = append_meetings_live(excel_path, meetings)
        except PermissionError:
            messagebox.showerror(APP_TITLE, "写入失败：请先关闭正在打开的 Excel 文件后再试。")
            return
        except RuntimeError:
            try:
                count, row = append_meetings(excel_path, meetings)
            except PermissionError:
                messagebox.showerror(APP_TITLE, "写入失败：Excel 文件可能被占用，请关闭正在打开的 Excel 文件后再试。")
                return
            except Exception as exc:
                messagebox.showerror(APP_TITLE, f"写入失败：\n{exc}")
                return
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"写入失败：\n{exc}")
            return

        if count == 1:
            messagebox.showinfo(APP_TITLE, f"已添加 1 条会议，排序后位于第 {row} 行。")
        else:
            messagebox.showinfo(APP_TITLE, f"已添加 {count} 条会议，并已完成排序和格式整理。")

    def clear(self) -> None:
        self.input_text.delete("1.0", END)
        self.preview_text.configure(state="normal")
        self.preview_text.delete("1.0", END)
        self.preview_text.configure(state="disabled")


if __name__ == "__main__":
    root = Tk()
    app = MeetingAssistantApp(root)
    root.mainloop()
